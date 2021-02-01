import openpyxl
import pprint
import re
import random
# 問題文が複数行に渡り記述してある
# 解答が問題文と列がずれているものがある
# 正解が太字になっていないものがある
# 問題文と解答には一行空白がある（逆もしかり）

# ?what is class Excel → excel file get text


class Execl:
    # ToDo cell自動取得によるテキスト抽出
    def __init__(self, file_path):
        self.excel_obj = None
        self.file_path = file_path
        self.excel_path = None
        self.sheet_range = [["C1", "D139"], ["A1", "B1310"]]
        # self.sheet_range = []
        self.text = []
        self.sheet_count = None

    def _GetFileExcelPath(self):
        with open(self.file_path, mode="r", encoding="utf-8") as f:
            self.excel_path = f.readlines()[0]

    def _GetExcelFile(self):
        self.excel_obj = openpyxl.load_workbook(self.excel_path)
        #  cell番号自動取得
        # for i in range(len(self.excel_obj.sheetnames)):
        #     temp = [self.excel_obj.worksheets[i].max_row,
        #             self.excel_obj.worksheets[i].max_column]
        # self.sheet_range.append(temp)
        self.sheet_count = len(self.excel_obj.sheetnames)

    def _ExcelGetCell(self, index=0):
        sheet = self.excel_obj.worksheets[index]
        cell = sheet[self.sheet_range[index][0]:self.sheet_range[index][1]]
        for i, (c, d) in enumerate(cell):
            if c.value is not None or d.value is not None:
                tc = c.value if c.value is not None else ""
                td = d.value if d.value is not None else ""
                set_cell = c if c.value is not None else d
                self.text.append([i, tc + td, set_cell])
        # print(len(self.text))

# ?what is class Question → question management


class Question:
    def __init__(self):
        self.no = None
        self.question = ""
        self.choices = []
        self.answer = []
        self.row = None
        self.col = None
        self.choices_flag = False  # 正解が設定されているか判定

    def _SetQuestion(self, text):
        if text is not None:
            self.question += text

    def _SetAnswer(self, text):
        if text is not None:
            self.choices.append(text)

    def _SetProblem(self, text):
        if text is not None:
            self.answer.append(text)

    def _SetRow(self, row):
        if row is not None:
            self.row = row

    def _SetCol(self, col):
        if col is not None:
            self.col = col

    def _SetNo(self, no):
        if no is not None:
            self.no = no

    def _CheckAnswerCount(self):
        if len(self.answer) >= 1:
            self.choices_flag = True
        else:
            self.choices_flag = False

    def _CheckBold(self, cell, text):
        flag = cell.font.bold
        flag = False if flag is None else flag
        if flag:
            self._SetProblem(text)

    def _QuestionResultPrint(self):
        print("-" * 10)
        print(self.no + 1, "問題目")
        print(self.question)
        # pprint.pprint(self.choices)
        print("選択肢")
        for i in self.choices:
            print(i)
        print("正解")
        for i in self.answer:
            print(i)
        print("-" * 10)

    def _QuestionPrint(self):
        self._HaihunPrint(str(self.no + 1) + "問")
        print(self.question)
        # print("\n")

    def _AnswerPrint(self):
        self._HaihunPrint("選択肢")
        for i in self.choices:
            print(i.replace("\n", ""))
        self._HaihunPrint("解答入力")

    def _ProblemPrint(self):
        self._HaihunPrint("正解")
        for i in self.answer:
            print(i)
        self._HaihunPrint("")

    def _HaihunPrint(self, text="問題"):
        # print("\n")
        temp = "-" * 5
        print(temp, text, temp)


# ?what is class AWSCloudPractitioner → aws management

class AWSCloudPractitioner:

    def __init__(self, text_list):
        self.untreated_text_list = text_list
        self.question_list = []
        # 出題問題オブジェクト、選択解答
        self.setting_question = []
        self.mode = True
        self.question_Definition = {"q": "", "a": "^[a-zA-Z][.]"}
        self.number_of_correct_answers = 0
        self.print_mode = 0
        # self.reorganization_list = None

    def _CheckAnswerSreach(self):
        # print(len(self.question_list))
        # print(self.question_list[209]._QuestionResultPrint())
        for i in self.question_list:
            if not i.choices_flag:
                i._QuestionResultPrint()

    def _ReorganizationQuestionList(self):
        # question_flag = False
        add_flag = False
        answer_flag = False
        q = Question()
        for i in self.untreated_text_list:
            # 問題文判定
            if not self._CheckProblemText(i[1]):
                if answer_flag:
                    add_flag = True
                else:
                    q.question += i[1]
            else:
                answer_flag = True
                q._CheckBold(i[2], i[1])
                q.choices.append(i[1])
            if add_flag:
                add_flag = False
                answer_flag = False
                q._SetNo(len(self.question_list))
                q._CheckAnswerCount()
                self.question_list.append(q)
                q = Question()
                q.question += i[1]

    def _CheckProblemText(self, text, juge="a"):
        pattern = self.question_Definition[juge]
        matchOB = re.match(pattern, text.strip())
        if matchOB:
            return True
        else:
            return False

    def _MockExamination(self, count=10):
        for i in range(count):
            temp = random.choice(self.question_list)
            select = {"obj": temp, "choices": None}
            if temp.choices_flag:
                temp._QuestionPrint()
                temp._AnswerPrint()
                # temp._ProblemPrint()
                select["choices"] = self._InputUserAnswer()
                self._PrintAnswerJuge(temp)
            else:
                print("answer not set error")
            self.setting_question.append(select)
        self._CheckingAnswers()

    def _InputUserAnswer(self):
        print("解答を選択してください")
        print("a, A, 1 のどれでも可")
        print("複数選択肢は「,」、「.」のどちらかで")
        user_answer = input()
        while True:
            if len(user_answer) > 0:
                break
            else:
                print("選択肢を入力してください。")
            user_answer = input()
        return user_answer

    def _PrintAnswerJuge(self, question):
        if self.mode:
            question._ProblemPrint()

    def _CheckingAnswers(self):
        for i in self.setting_question:
            answer_list = i["obj"].answer
            answer = self._ReplaceAnswerList(answer_list)
            choice = i["choices"].replace(".", "").replace(",", "")
            choice = list(self._ReplaceAlphabetString(choice))
            answer = self._SortListAlphabet(answer)
            choice = self._SortListAlphabet(choice)
            if answer == choice:
                self.number_of_correct_answers += 1
        self._QuestionResult()

    def _QuestionResult(self):
        self.setting_question[0]["obj"]._HaihunPrint("結果")
        print(str(self.number_of_correct_answers) +
              "/" + str(len(self.setting_question)))

    def _ReplaceAnswerList(self, answer_list):
        result = ""
        for i in answer_list:
            result += i[0]
        return list(result)

    def _ReplaceAlphabetString(self, text):
        alphabet = [chr(i) for i in range(65, 65 + 26)]
        number = list(map(str, list(range(1, 10))))
        d = dict(zip(number, alphabet))
        # print(s.translate(str.maketrans(d)))
        return text.translate(str.maketrans(d))

    def _SortListAlphabet(self, sort_list):
        return sort_list.sort()


def cd():
    import os
    os.chdir(os.path.dirname(__file__))


def main():
    ex = Execl("../mine/path.txt")
    ex._GetFileExcelPath()
    ex._GetExcelFile()
    for i in range(ex.sheet_count):
        ex._ExcelGetCell(i)
    aws = AWSCloudPractitioner(ex.text)
    aws._ReorganizationQuestionList()
    # print(aws.question_list[0])
    # print(len(aws.question_list))
    # for i in range(10):
    #     print(i, "番目", ex.text[i])
    # aws._MockExamination(1)
    print(aws.question_list[210].choices_flag)
    aws._CheckAnswerSreach()
    # for i in range(3):
    #     temp = random.choice(aws.question_list)
    #     temp._QuestionPrint()
    #     temp._AnswerPrint()
    #     temp._ProblemPrint()


def test():
    ex = Execl("../mine/path.txt")
    ex._GetFileExcelPath()
    ex._GetExcelFile()
    # for i in range(ex.sheet_count):
    ex._ExcelGetCell(0)
    aws = AWSCloudPractitioner(ex.text)
    aws._ReorganizationQuestionList()
    for i in aws.question_list:
        pprint.pprint(i._QuestionResultPrint())
    # pprint.pprint(aws.question_list._QuestionResultPrint())


if __name__ == "__main__":
    cd()
    test()
    # main()
    # pprint.pprint(d)
    # print(alphabet, number)
